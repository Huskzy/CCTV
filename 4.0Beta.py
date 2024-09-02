from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import base64
import os
import time
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

def sanitize_filename(filename):
    return re.sub(r'[^\w\-]', '', filename)

def save_canvas_as_image(canvas, button_text, base_folder):
    data_url = driver.execute_script('return arguments[0].toDataURL("image/png").substring(22);', canvas)
    image_data = base64.b64decode(data_url)
    sanitized_name = sanitize_filename(button_text)
    image_path = os.path.join(base_folder, f'{sanitized_name}.png')
    os.makedirs(os.path.dirname(image_path), exist_ok=True)
    with open(image_path, 'wb') as file:
        file.write(image_data)
    print(f"已保存圖片 {image_path}")
    return image_path

def print_creator_info(version):
    print("Jim Chen")
    print(f"Ver：{version}")
    print("mail：x25687441@gmail.com")
    print("---------------")

def wait_for_canvas_and_save(timeout=10, initial_wait=3, post_save_wait=2):
    start_time = time.time()
    end_time = start_time + timeout
    time.sleep(initial_wait)

    while time.time() < end_time:
        try:
            canvas = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'canvas'))
            )
            if canvas:
                image_path = save_canvas_as_image(canvas, sanitized_text, base_folder)
                time.sleep(post_save_wait)
                return image_path
        except Exception as e:
            print(f"等待 canvas 元素出錯: {e}")
            time.sleep(1)
    print("超過等待時間，無法擷取圖片")
    return None

def get_file_size_in_kb(file_path):
    return os.path.getsize(file_path) / 1024

def extract_characters_in_range(file_path):
    # 提取文件名（不包括扩展名）
    file_name = os.path.basename(file_path)
    
    # 找到文件名中所有的中文字符
    chinese_characters = re.findall(r'[\u4e00-\u9fff]', file_name)
    
    if not chinese_characters:
        return ''

    # 找到第一个和最后一个中文字符的位置
    first_chinese_index = file_name.find(chinese_characters[0])
    last_chinese_index = file_name.rfind(chinese_characters[-1]) + 1

    # 提取从第一个到最后一个中文字符之间的所有字符
    extracted_name = file_name[first_chinese_index:last_chinese_index]
    
    return extracted_name

def save_image_records_to_excel(image_records, base_folder):
    try:
        df = pd.DataFrame(image_records)
        excel_path = os.path.join(base_folder, 'image_records.xlsx')
        
        # 创建 Excel 文件并写入数据
        wb = Workbook()
        ws = wb.active

        # 写入表头
        headers = ['File ID', 'Image Path', 'Status']
        ws.append(headers)

        # 设置列宽
        ws.column_dimensions['A'].width = 10.5  # File ID 列宽
        ws.column_dimensions['B'].width = 26    # Image Path 列宽
        ws.column_dimensions['C'].width = 6  # Status 列宽

        # 添加数据和状态
        for record in image_records:
            file_id = record['File ID']
            image_path = record['Image Path']
            cleaned_name = extract_characters_in_range(image_path)
            file_size_kb = get_file_size_in_kb(image_path)
            status = '離線' if file_size_kb <= 2 else '上線'
            
            # 写入数据
            ws.append([file_id, cleaned_name, status])
            # 设置单元格超链接
            relative_path = os.path.relpath(image_path, start=base_folder)
            ws.cell(row=ws.max_row, column=2).hyperlink = relative_path
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(horizontal='left')

        wb.save(excel_path)
        print(f"圖片記錄已保存到 {excel_path}")
    except Exception as e:
        print(f"保存 Excel 文件時出錯: {e}")


# 获取当前时间，并创建时间命名的文件夹
current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
base_folder = os.path.join('canvas_images', current_time)

os.makedirs(base_folder, exist_ok=True)

version = os.path.basename(__file__).replace('.py', '')

print_creator_info(version)

chrome_options = Options()
chrome_options.add_argument("--disable-notifications")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

image_records = []

try:
    driver.get('http://125.227.199.196:85/login')

    wait = WebDriverWait(driver, 10)

    username_input = driver.find_element(By.NAME, 'username')
    password_input = driver.find_element(By.NAME, 'password')

    username_input.send_keys('kris')
    password_input.send_keys('89723649')

    login_button = driver.find_element(By.XPATH, '//button[contains(span, "Login")]')
    login_button.click()

    wait.until(EC.url_changes('http://125.227.199.196:85/login'))
    print("登錄成功。")

    liveview_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//a[span[text()="LiveView"]]')))
    liveview_button.click()

    time.sleep(3)

    processed_buttons = set()

    while True:
        buttons = driver.find_elements(By.CSS_SELECTOR, 'button.list-group-item-action')

        if not buttons:
            print("無更多按鈕。")
            break

        for button in buttons:
            try:
                button_text = button.text.strip()

                if button_text in processed_buttons:
                    continue

                sanitized_text = sanitize_filename(button_text)

                button.click()

                image_path = wait_for_canvas_and_save(timeout=10, initial_wait=3, post_save_wait=2)

                if image_path:
                    processed_buttons.add(button_text)
                    file_id = sanitized_text[:8]
                    image_record = {'File ID': file_id, 'Image Path': image_path}
                    image_records.append(image_record)
                    save_image_records_to_excel(image_records, base_folder)  # 立即保存 Excel 文件

                driver.back()
                time.sleep(1)

            except Exception as e:
                print(f"處理按鈕 {button_text} 時出錯: {e}")

finally:
    driver.quit()

print("處理完成。")